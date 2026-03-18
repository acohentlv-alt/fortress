"""Client API routes — Smart Upload Engine for CSV/XLSX ingestion.

Intelligently maps ANY column layout to Fortress schema using the
alias dictionary in column_mapper.py. Supports:
- CSV (comma, semicolon, tab delimiters)
- XLSX (via openpyxl)
- SIRET/SIREN/TVA intracommunautaire normalization
- Full data ingestion: companies, contacts, officers
- Overflow columns → extra_data JSONB

Each upload creates a scrape_jobs entry (mode='upload') for tracking.
"""

from __future__ import annotations

import csv
import io
import json
import logging
import re
import traceback
from datetime import datetime, timezone

from fastapi import APIRouter, File, UploadFile
from fastapi.responses import JSONResponse

from psycopg.types.json import Json

from fortress.api.column_mapper import MappingResult, map_columns, normalize_siren
from fortress.api.db import fetch_all, fetch_one, get_conn
from fortress.api.routes.activity import log_activity

router = APIRouter(prefix="/api/client", tags=["client"])
logger = logging.getLogger("fortress.api.client")


# ---------------------------------------------------------------------------
# Upload endpoint — smart ingestion
# ---------------------------------------------------------------------------

@router.post("/upload")
async def upload_client_file(file: UploadFile = File(...)):
    """Upload a CSV or XLSX file. Intelligently maps columns and ingests data."""
    content = await file.read()
    filename = file.filename or "upload"
    is_xlsx = filename.lower().endswith((".xlsx", ".xls"))

    # ── Step 1: Parse file into headers + rows ──
    try:
        if is_xlsx:
            headers, rows = _parse_xlsx(content)
        else:
            headers, rows = _parse_csv(content)
    except Exception as e:
        logger.error("Parse error for %s: %s", filename, e)
        return JSONResponse(status_code=400, content={
            "error": f"Erreur de lecture: {str(e)}"
        })

    if not headers or not rows:
        return JSONResponse(status_code=400, content={
            "error": "Fichier vide ou illisible."
        })

    # ── Step 2: Map columns ──
    mapping = map_columns(headers)

    if mapping.siren_column is None:
        return JSONResponse(status_code=400, content={
            "error": "Aucune colonne SIREN/SIRET détectée dans le fichier.",
            "headers": headers[:20],
        })

    # ── Step 3: Create upload job ──
    now = datetime.now(tz=timezone.utc)
    query_id = f"upload_{re.sub(r'[^a-zA-Z0-9]', '_', filename)[:50]}_{int(now.timestamp())}"

    try:
        async with get_conn() as conn:
            await conn.execute("""
                INSERT INTO scrape_jobs
                    (query_id, query_name, status, batch_size, total_companies,
                     strategy, mode, created_at, updated_at)
                VALUES (%s, %s, 'in_progress', %s, %s, 'upload', 'upload', %s, %s)
            """, (query_id, f"Import: {filename}", len(rows), len(rows), now, now))
    except RuntimeError as exc:
        return JSONResponse(status_code=503, content={"error": str(exc)})
    except Exception as exc:
        logger.error("Failed to create upload job: %s", exc)
        return JSONResponse(status_code=500, content={"error": f"Erreur DB: {exc}"})

    # ── Step 4: Ingest data ──
    stats = {
        "companies_inserted": 0,
        "companies_updated": 0,
        "contacts_upserted": 0,
        "officers_upserted": 0,
        "rows_skipped": 0,
        "siren_invalid": 0,
        "errors": [],
    }

    try:
        async with get_conn() as conn:
            for row_idx, row in enumerate(rows):
                try:
                    # Savepoint per row: if one row fails, only that
                    # savepoint rolls back — the outer transaction stays clean.
                    async with conn.transaction():
                        await _ingest_row(conn, row, headers, mapping, stats, query_id)
                except Exception as exc:
                    stats["rows_skipped"] += 1
                    err_msg = f"Row {row_idx + 2}: {type(exc).__name__}: {exc}"
                    logger.warning("Ingest error: %s", err_msg)
                    if len(stats["errors"]) < 20:
                        stats["errors"].append(err_msg)

                # Update progress every 50 rows
                if (row_idx + 1) % 50 == 0:
                    scraped = stats["companies_inserted"] + stats["companies_updated"]
                    async with conn.transaction():
                        await conn.execute("""
                            UPDATE scrape_jobs
                            SET companies_scraped = %s, updated_at = %s
                            WHERE query_id = %s
                        """, (scraped, datetime.now(tz=timezone.utc), query_id))

            # Tag all ingested companies in query_tags
            async with conn.transaction():
                await conn.execute("""
                    INSERT INTO query_tags (siren, query_name, tagged_at)
                    SELECT DISTINCT sa.siren, %s, NOW()
                    FROM scrape_audit sa
                    WHERE sa.query_id = %s
                    ON CONFLICT (siren, query_name) DO NOTHING
                """, (query_id, query_id))

            # Mark job completed
            scraped = stats["companies_inserted"] + stats["companies_updated"]
            async with conn.transaction():
                await conn.execute("""
                    UPDATE scrape_jobs
                    SET status = 'completed',
                        companies_scraped = %s,
                        companies_qualified = %s,
                        updated_at = %s
                    WHERE query_id = %s
                """, (scraped, stats["contacts_upserted"], datetime.now(tz=timezone.utc), query_id))

    except RuntimeError as exc:
        return JSONResponse(status_code=503, content={"error": str(exc)})
    except Exception as exc:
        logger.error("Ingestion failed: %s\n%s", exc, traceback.format_exc())
        # Still try to mark job as failed
        try:
            async with get_conn() as conn:
                await conn.execute("""
                    UPDATE scrape_jobs SET status = 'failed', updated_at = %s
                    WHERE query_id = %s
                """, (datetime.now(tz=timezone.utc), query_id))
        except Exception:
            pass
        return JSONResponse(status_code=500, content={
            "error": f"Erreur d'ingestion: {exc}",
            "stats": {k: v for k, v in stats.items() if k != "errors"},
        })

    # ── Step 5: Build mapping summary for frontend ──
    mapping_summary = _build_mapping_summary(mapping)

    # Keep first 5 errors for diagnosability, strip the rest
    errors = stats.pop("errors", [])
    error_count = len(errors)

    # Log activity
    scraped_total = stats["companies_inserted"] + stats["companies_updated"]
    await log_activity(
        user_id=None,
        username='system',
        action='upload',
        target_type='upload',
        target_id=query_id,
        details=f"Import {filename} — {scraped_total} entreprises, {stats['contacts_upserted']} contacts",
    )

    return {
        "status": "ok",
        "filename": filename,
        "query_id": query_id,
        "total_rows": len(rows),
        "stats": stats,
        "error_count": error_count,
        "error_samples": errors[:5],
        "mapping": mapping_summary,
    }


# ---------------------------------------------------------------------------
# Preview endpoint — show column mapping without ingesting
# ---------------------------------------------------------------------------

@router.post("/preview")
async def preview_upload(file: UploadFile = File(...)):
    """Parse file and return column mapping preview without ingesting."""
    content = await file.read()
    filename = file.filename or "upload"
    is_xlsx = filename.lower().endswith((".xlsx", ".xls"))

    try:
        if is_xlsx:
            headers, rows = _parse_xlsx(content)
        else:
            headers, rows = _parse_csv(content)
    except Exception as e:
        return JSONResponse(status_code=400, content={
            "error": f"Erreur de lecture: {str(e)}"
        })

    if not headers:
        return JSONResponse(status_code=400, content={
            "error": "Fichier vide ou illisible."
        })

    mapping = map_columns(headers)

    # Count valid SIRENs
    siren_count = 0
    unique_sirens = set()
    if mapping.siren_column is not None:
        for row in rows:
            siren_col = mapping.siren_column
            if siren_col < len(row):
                s = normalize_siren(str(row[siren_col] or ""))
                if s:
                    unique_sirens.add(s)
        siren_count = len(unique_sirens)

    return {
        "filename": filename,
        "total_rows": len(rows),
        "total_columns": len(headers),
        "valid_sirens": siren_count,
        "has_siren_column": mapping.siren_column is not None,
        "has_officer_data": mapping.has_officer_data,
        "mapping": _build_mapping_summary(mapping),
    }


# ---------------------------------------------------------------------------
# Stats endpoint
# ---------------------------------------------------------------------------

@router.get("/stats")
async def client_stats():
    """Return upload history and stats."""
    try:
        uploads_raw = await fetch_all("""
            SELECT sj.query_id, sj.query_name, sj.status, sj.batch_size,
                   sj.companies_scraped, sj.companies_qualified,
                   sj.created_at, sj.updated_at,
                   COUNT(DISTINCT sa.siren) AS siren_count
            FROM scrape_jobs sj
            LEFT JOIN scrape_audit sa ON sa.query_id = sj.query_id
            WHERE sj.mode = 'upload'
            GROUP BY sj.query_id, sj.query_name, sj.status, sj.batch_size,
                     sj.companies_scraped, sj.companies_qualified,
                     sj.created_at, sj.updated_at
            ORDER BY sj.created_at DESC
            LIMIT 20
        """)

        # Count total unique SIRENs across all upload jobs
        total_row = await fetch_one("""
            SELECT COUNT(DISTINCT sa.siren) AS total
            FROM scrape_audit sa
            JOIN scrape_jobs sj ON sj.query_id = sa.query_id
            WHERE sj.mode = 'upload'
        """)
        total_sirens = (total_row or {}).get("total", 0)

    except RuntimeError as exc:
        return JSONResponse(status_code=503, content={"error": str(exc)})

    # Reshape for frontend: extract filename from query_name "Import: filename"
    uploads = []
    for u in (uploads_raw or []):
        qn = u.get("query_name", "")
        source_file = qn.replace("Import: ", "") if qn.startswith("Import: ") else qn
        uploads.append({
            **u,
            "source_file": source_file,
            "siren_count": u.get("siren_count", 0),
            "uploaded_at": u.get("created_at"),
        })

    return {
        "uploads": uploads,
        "total_sirens": total_sirens,
    }


@router.delete("/clear")
async def clear_client_sirens():
    """Clear all client-uploaded data (upload-mode jobs and their query_tags)."""
    try:
        async with get_conn() as conn:
            # Get upload job query_names to clear their tags
            jobs = await conn.execute(
                "SELECT query_name FROM scrape_jobs WHERE mode = 'upload'"
            )
            job_rows = await jobs.fetchall()
            query_names = [r[0] for r in job_rows] if job_rows else []

            deleted_tags = 0
            for qn in query_names:
                res = await conn.execute(
                    "DELETE FROM query_tags WHERE query_name = %s", (qn,)
                )
                deleted_tags += res.rowcount

            # Delete the upload jobs themselves
            res = await conn.execute(
                "DELETE FROM scrape_jobs WHERE mode = 'upload'"
            )
            deleted_jobs = res.rowcount
            await conn.commit()

        return {
            "cleared": True,
            "deleted_jobs": deleted_jobs,
            "deleted_tags": deleted_tags,
        }
    except RuntimeError as exc:
        return JSONResponse(status_code=503, content={"error": str(exc)})


# ---------------------------------------------------------------------------
# Internal: Row ingestion
# ---------------------------------------------------------------------------

async def _ingest_row(
    conn, row: list, headers: list[str],
    mapping: MappingResult, stats: dict, query_id: str,
):
    """Ingest a single row into companies, contacts, and officers tables."""

    # Extract SIREN
    siren_col = mapping.siren_column
    if siren_col is None or siren_col >= len(row):
        stats["siren_invalid"] += 1
        return

    raw_siren = str(row[siren_col] or "").strip()
    siren = normalize_siren(raw_siren)
    if not siren:
        stats["siren_invalid"] += 1
        return

    # Organize values by target table
    company_fields: dict[str, str] = {}
    contact_fields: dict[str, str] = {}
    officer_fields: dict[str, str] = {}
    extra_data: dict[str, str] = {}

    for col_idx, col_mapping in enumerate(mapping.columns):
        if col_idx >= len(row):
            break
        value = str(row[col_idx] or "").strip() if row[col_idx] else ""
        if not value or value == "-":
            continue

        if col_mapping.target_table == "companies":
            if col_mapping.target_field not in company_fields:
                company_fields[col_mapping.target_field] = value
        elif col_mapping.target_table == "contacts":
            contact_fields[col_mapping.target_field] = value
        elif col_mapping.target_table == "officers":
            officer_fields[col_mapping.target_field] = value
        elif col_mapping.target_table == "extra_data":
            extra_data[col_mapping.target_field] = value
        # skip → ignore

    # ── Upsert company ──
    await _upsert_company(conn, siren, company_fields, extra_data, stats)

    # ── Upsert contacts ──
    if any(f in contact_fields for f in ("phone", "email", "website")):
        await _upsert_contact(conn, siren, contact_fields, stats)

    # ── Upsert officer ──
    if officer_fields.get("nom") or officer_fields.get("prenom"):
        await _upsert_officer(conn, siren, officer_fields, stats)

    # ── Audit trail ──
    await conn.execute("""
        INSERT INTO scrape_audit (query_id, siren, action, result, timestamp)
        VALUES (%s, %s, 'upload', 'success', %s)
    """, (query_id, siren, datetime.now(tz=timezone.utc)))


async def _upsert_company(conn, siren: str, fields: dict, extra: dict, stats: dict):
    """Insert or update a company record.

    For existing SIRENE companies: fills empty fields only (COALESCE).
    For new companies: INSERT with ON CONFLICT fall-through.
    """
    # Check if company exists
    cur = await conn.execute(
        "SELECT siren FROM companies WHERE siren = %s", (siren,)
    )
    row = await cur.fetchone()

    if row:
        # Update existing: only fill empty fields
        sets = []
        vals = []
        for db_field, value in fields.items():
            if db_field == "siren":
                continue
            value = _cast_field(db_field, value)
            if value is None:
                continue
            sets.append(f"{db_field} = COALESCE({db_field}, %s)")
            vals.append(value)

        if extra:
            # COALESCE handles NULL extra_data for existing SIRENE companies
            sets.append("extra_data = COALESCE(extra_data, '{}'::jsonb) || %s::jsonb")
            vals.append(Json(extra))

        if sets:
            sets.append("updated_at = %s")
            vals.append(datetime.now(tz=timezone.utc))
            vals.append(siren)
            await conn.execute(
                f"UPDATE companies SET {', '.join(sets)} WHERE siren = %s",
                tuple(vals),
            )
        stats["companies_updated"] += 1
    else:
        # Insert new company (not in SIRENE base)
        denomination = fields.get("denomination", f"Entreprise {siren}")
        insert_fields = {"siren": siren, "denomination": denomination, "statut": "A"}
        for db_field, value in fields.items():
            if db_field in ("siren", "denomination"):
                continue
            casted = _cast_field(db_field, value)
            if casted is not None:
                insert_fields[db_field] = casted

        if extra:
            insert_fields["extra_data"] = Json(extra)

        cols = ", ".join(insert_fields.keys())
        placeholders = ", ".join(["%s"] * len(insert_fields))

        # ON CONFLICT: if SIREN was inserted between our SELECT and INSERT
        update_clauses = []
        for k in insert_fields:
            if k not in ("siren", "statut"):
                update_clauses.append(
                    f"{k} = COALESCE(companies.{k}, EXCLUDED.{k})"
                )
        update_clauses.append("updated_at = NOW()")

        await conn.execute(
            f"INSERT INTO companies ({cols}) VALUES ({placeholders})"
            f" ON CONFLICT (siren) DO UPDATE SET {', '.join(update_clauses)}",
            tuple(insert_fields.values()),
        )
        stats["companies_inserted"] += 1


async def _upsert_contact(conn, siren: str, fields: dict, stats: dict):
    """Upsert a contact row with source='upload'."""
    cols = ["siren", "source"]
    vals: list = [siren, "upload"]

    for db_field in ("phone", "email", "website", "social_linkedin",
                     "social_facebook", "social_twitter"):
        if db_field in fields:
            cols.append(db_field)
            vals.append(fields[db_field])

    col_str = ", ".join(cols)
    placeholders = ", ".join(["%s"] * len(cols))

    updates = []
    for db_field in cols:
        if db_field in ("siren", "source"):
            continue
        updates.append(f"{db_field} = COALESCE(contacts.{db_field}, EXCLUDED.{db_field})")

    if not updates:
        return

    await conn.execute(
        f"INSERT INTO contacts ({col_str}) VALUES ({placeholders})"
        f" ON CONFLICT (siren, source) DO UPDATE SET {', '.join(updates)}",
        tuple(vals),
    )
    stats["contacts_upserted"] += 1


async def _upsert_officer(conn, siren: str, fields: dict, stats: dict):
    """Upsert an officer record."""
    nom = fields.get("nom", "")
    prenom = fields.get("prenom", "")

    if not nom and not prenom:
        return

    await conn.execute("""
        INSERT INTO officers (siren, nom, prenom, role, source,
                              civilite, email_direct, ligne_directe,
                              code_fonction, type_fonction)
        VALUES (%s, %s, %s, %s, 'upload', %s, %s, %s, %s, %s)
        ON CONFLICT (siren, nom, COALESCE(prenom, ''))
        DO UPDATE SET
            role = COALESCE(officers.role, EXCLUDED.role),
            civilite = COALESCE(officers.civilite, EXCLUDED.civilite),
            email_direct = COALESCE(officers.email_direct, EXCLUDED.email_direct),
            ligne_directe = COALESCE(officers.ligne_directe, EXCLUDED.ligne_directe),
            code_fonction = COALESCE(officers.code_fonction, EXCLUDED.code_fonction),
            type_fonction = COALESCE(officers.type_fonction, EXCLUDED.type_fonction)
    """, (
        siren, nom, prenom,
        fields.get("role"),
        fields.get("civilite"),
        fields.get("email_direct"),
        fields.get("ligne_directe"),
        fields.get("code_fonction"),
        fields.get("type_fonction"),
    ))
    stats["officers_upserted"] += 1


def _cast_field(db_field: str, value: str):
    """Cast a string value to the correct Python type for a DB field."""
    if db_field == "chiffre_affaires":
        try:
            return int(re.sub(r"[^\d]", "", str(value)))
        except (ValueError, TypeError):
            return None
    elif db_field == "annee_ca":
        try:
            return int(value)
        except (ValueError, TypeError):
            return None
    elif db_field == "date_fondation":
        # Try DD/MM/YYYY then YYYY
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%Y"):
            try:
                from datetime import date
                return date.strftime(datetime.strptime(str(value), fmt), "%Y-%m-%d")
            except (ValueError, TypeError):
                continue
        return None
    return value


# ---------------------------------------------------------------------------
# Internal: File parsers
# ---------------------------------------------------------------------------

def _parse_xlsx(content: bytes) -> tuple[list[str], list[list]]:
    """Parse XLSX bytes into headers + rows using openpyxl."""
    import openpyxl
    from io import BytesIO
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(cell or "") for cell in next(rows_iter)]
    rows = [list(row) for row in rows_iter if any(cell for cell in row)]
    wb.close()
    return headers, rows


def _parse_csv(content: bytes) -> tuple[list[str], list[list]]:
    """Parse CSV bytes into headers + rows. Auto-detects delimiter + encoding."""
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    first_line = text.split("\n")[0] if text else ""
    semicolons = first_line.count(";")
    commas = first_line.count(",")
    tabs = first_line.count("\t")
    delimiter = ";"  if semicolons > commas and semicolons > tabs \
        else "\t" if tabs > commas and tabs > semicolons \
        else ","

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    all_rows = [row for row in reader if any(cell.strip() for cell in row)]

    if not all_rows:
        return [], []

    headers = [cell.strip() for cell in all_rows[0]]
    rows = all_rows[1:]
    return headers, rows


# ---------------------------------------------------------------------------
# Internal: Mapping summary
# ---------------------------------------------------------------------------

def _build_mapping_summary(mapping: MappingResult) -> dict:
    """Build a JSON-friendly mapping summary for the frontend."""
    recognized = []
    overflow = []
    skipped = []

    for col in mapping.columns:
        entry = {
            "source": col.source_name,
            "target": f"{col.target_table}.{col.target_field}" if col.target_field else None,
            "confidence": col.confidence,
        }
        if col.target_table == "skip":
            skipped.append(entry)
        elif col.target_table == "extra_data":
            overflow.append(entry)
        else:
            recognized.append(entry)

    return {
        "recognized_count": len(recognized),
        "overflow_count": len(overflow),
        "skipped_count": len(skipped),
        "recognized": recognized,
        "overflow": overflow,
    }
