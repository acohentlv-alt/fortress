"""Export API routes — CSV, XLSX, and JSONL downloads."""

import csv
import decimal
import io
import json


class _DecimalEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, decimal.Decimal):
            return float(obj)
        return super().default(obj)

import re

from fastapi import APIRouter, Query, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from fortress.api.db import fetch_all, fetch_one
from fortress.api.sql_helpers import merged_contacts_cte

router = APIRouter(prefix="/api/export", tags=["export"])

# Map SIRENE numeric codes to human-readable legal form labels
_FORME_JURIDIQUE_LABELS = {
    '1000': 'Entrepreneur individuel', '5306': 'EURL', '5307': 'SA',
    '5370': 'SAS', '5498': 'EURL', '5499': 'SARL',
    '5505': 'SA', '5510': 'SAS', '5515': 'SNC',
    '5520': 'SCS', '5522': 'SCA', '5525': 'SARL unipersonnelle',
    '5530': 'SELASU', '5532': 'SELAS', '5560': 'SCI', '5599': 'SA',
    '5710': 'SAS', '5720': 'SASU', '9220': 'Association loi 1901',
    '9221': 'Association déclarée', '6316': 'SCOP', '6317': 'SCOP',
}


_LINK_CONFIDENCE_LABELS = {
    'confirmed': 'Confirmé',
    'pending': 'En attente',
}


def _fmt(col_key: str, value) -> str:
    """Format a cell value for export. Converts forme_juridique codes to labels
    and link_confidence codes to French labels."""
    if not value:
        return ""
    if col_key == "forme_juridique":
        return _FORME_JURIDIQUE_LABELS.get(str(value), str(value))
    if col_key == "link_confidence":
        return _LINK_CONFIDENCE_LABELS.get(str(value), "")
    return str(value)

_CSV_COLUMNS = [
    ("Nom", "denomination"),
    ("SIREN", "effective_siren"),
    ("SIRET", "siret_siege"),
    ("NAF", "naf_code"),
    ("Activité", "naf_libelle"),
    ("Forme juridique", "forme_juridique"),
    ("Adresse", "adresse"),
    ("Code postal", "code_postal"),
    ("Ville", "ville"),
    ("Département", "departement"),
    ("Statut SIRENE", "statut"),       # renamed from "Statut" to avoid clash with the new column
    ("Date création", "date_creation"),
    ("Effectif", "tranche_effectif"),
    ("Téléphone", "phone"),
    ("Email", "email"),
    ("Site web", "website"),
    ("Adresse Maps", "address"),
    ("Google Maps URL", "maps_url"),
    ("LinkedIn", "social_linkedin"),
    ("Facebook", "social_facebook"),
    ("Twitter", "social_twitter"),
    ("Note Google", "rating"),
    ("Avis Google", "review_count"),
    ("Notes", "notes"),
    ("Statut lien", "link_confidence"),  # NEW: Confirmé / En attente / blank
]

# SELECT columns used by every export query. Computes the REAL SIREN:
# - Pure SIRENE entities: use co.siren directly
# - MAPS entities with confirmed link: use linked_siren (real SIREN)
# - MAPS entities with pending link: use linked_siren (candidate SIREN, flagged via Statut lien)
# - MAPS entities without confirmed link: NULL (avoids exporting MAPS00001 as "SIREN")
#
# For pending MAPS rows (Apr 29), SIRENE-side reference fields (SIRET, NAF, forme
# juridique, statut, date création, effectif) are blank because
# `_copy_sirene_reference_data` only runs on auto-confirm or manual /link approve.
# To show the entity "as if merged" — so Cindy gets the candidate's full data to
# verify — we LEFT JOIN companies via `linked_siren` (alias `sirene_ref`) and
# COALESCE the SIRENE-side fields. Maps-derived fields (denomination, adresse,
# code_postal, ville) are kept as-is to preserve storefront location accuracy
# (Frankenstein fix Apr 22).
_EXPORT_SELECT = """
    CASE
        WHEN co.siren NOT LIKE 'MAPS%%' THEN co.siren
        WHEN co.link_confidence = 'confirmed' THEN co.linked_siren
        WHEN co.link_confidence = 'pending' THEN co.linked_siren
        ELSE NULL
    END AS effective_siren,
    COALESCE(co.siret_siege, sirene_ref.siret_siege) AS siret_siege,
    co.denomination,
    COALESCE(co.naf_code, sirene_ref.naf_code) AS naf_code,
    COALESCE(co.naf_libelle, sirene_ref.naf_libelle) AS naf_libelle,
    COALESCE(co.forme_juridique, sirene_ref.forme_juridique) AS forme_juridique,
    co.adresse,
    COALESCE(co.code_postal, sirene_ref.code_postal) AS code_postal,
    COALESCE(co.ville, sirene_ref.ville) AS ville,
    co.departement,
    COALESCE(co.statut, sirene_ref.statut) AS statut,
    COALESCE(co.date_creation, sirene_ref.date_creation) AS date_creation,
    COALESCE(co.tranche_effectif, sirene_ref.tranche_effectif) AS tranche_effectif,
    mc.phone, mc.email, mc.website,
    mc.address, mc.maps_url,
    mc.social_linkedin, mc.social_facebook, mc.social_twitter,
    mc.rating, mc.review_count,
    cn.notes,
    co.link_confidence
"""

# JOINs used by every export query.
# `sirene_ref` join hydrates SIRENE-side fields for MAPS entities (both confirmed
# and pending). For pure SIRENE rows (co.siren NOT LIKE 'MAPS%'), this LEFT JOIN
# never matches (linked_siren IS NULL on pure SIRENE) — no impact.
_EXPORT_JOINS = """
    LEFT JOIN merged_contacts mc ON mc.siren = co.siren
    LEFT JOIN companies sirene_ref ON sirene_ref.siren = co.linked_siren
                                  AND co.siren LIKE 'MAPS%%'
    LEFT JOIN (
        SELECT siren, STRING_AGG(text, ' | ' ORDER BY created_at DESC) AS notes
        FROM company_notes
        GROUP BY siren
    ) cn ON cn.siren = co.siren
"""


async def _fetch_export_data(batch_id: str, search_query: str | None = None) -> list[dict]:
    """Fetch all companies + contacts for a specific batch.

    Scoped via batch_tags (one row per actual Maps result, keyed by
    batch_id). Earlier we used batch_log, but batch_log also stores A2
    candidate-lookup audit rows under each candidate's real SIREN —
    that polluted exports with empty ghost rows.

    Returns confirmed AND pending MAPS rows. Unmatched MAPS (no linked
    SIREN) are excluded — they have no real SIREN and would pollute the
    export. Confirmed rows are sorted first, pending rows last, so Cindy
    can act on confirmed contacts immediately and review pending ones
    afterwards. Pending rows are flagged via the `Statut lien` column
    ("En attente") and shaded in XLSX exports.

    When search_query is set (E4.A), only entities whose batch_log row
    recorded this exact search_query are returned.
    """
    job = await fetch_one(
        "SELECT batch_id FROM batch_data WHERE batch_id = %s", (batch_id,)
    )
    if not job:
        return []

    sq_clause = ""
    sq_params: tuple = ()
    if search_query:
        # E4.A — only entities whose batch_log row recorded this exact search_query.
        sq_clause = """
          AND co.siren IN (
              SELECT DISTINCT siren FROM batch_log
              WHERE batch_id = %s AND search_query = %s
          )
        """
        sq_params = (batch_id, search_query)

    return await fetch_all(f"""
        WITH {merged_contacts_cte('SELECT DISTINCT siren FROM batch_tags WHERE batch_id = %s')}
        SELECT {_EXPORT_SELECT}
        FROM (SELECT DISTINCT siren FROM batch_tags WHERE batch_id = %s) sa
        JOIN companies co ON co.siren = sa.siren
        {_EXPORT_JOINS}
        WHERE (co.siren NOT LIKE 'MAPS%%' OR co.link_confidence IN ('confirmed', 'pending'))
          AND (co.naf_status IS DISTINCT FROM 'mismatch' OR co.link_method IN ('chain', 'gemini_judge', 'siret_address_naf'))
          {sq_clause}
        ORDER BY (CASE WHEN co.link_confidence = 'pending' THEN 1 ELSE 0 END),
                 co.denomination
    """, (batch_id, batch_id) + sq_params)


# ── MASTER EXPORT (must be declared BEFORE parameterised routes) ──


@router.get("/master/csv")
async def export_master_csv(request: Request):
    """Download all SCRAPED companies across all queries as CSV.

    Admin only. Scoped to batch_tags (only companies we've actually processed),
    not the full 14.7M SIRENE table.
    """
    user = getattr(request.state, 'user', None)
    if not user or user.role != 'admin':
        from fastapi.responses import JSONResponse
        return JSONResponse(status_code=403, content={"error": "Admin uniquement"})
    rows = await fetch_all(f"""
        WITH {merged_contacts_cte('SELECT DISTINCT siren FROM batch_tags')}
        SELECT {_EXPORT_SELECT}
        FROM (SELECT DISTINCT siren FROM batch_tags) qt
        JOIN companies co ON co.siren = qt.siren
        {_EXPORT_JOINS}
        WHERE (co.siren NOT LIKE 'MAPS%%' OR co.link_confidence IN ('confirmed', 'pending'))
          AND (co.naf_status IS DISTINCT FROM 'mismatch' OR co.link_method IN ('chain', 'gemini_judge', 'siret_address_naf'))
        ORDER BY (CASE WHEN co.link_confidence = 'pending' THEN 1 ELSE 0 END),
                 co.denomination
    """)
    if not rows:
        return StreamingResponse(
            io.BytesIO(b"No data"),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=fortress_master.csv"},
        )

    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")
    writer.writerow([col[0] for col in _CSV_COLUMNS])
    for row in rows:
        writer.writerow([
            _fmt(col[1], row.get(col[1])) for col in _CSV_COLUMNS
        ])

    content = buf.getvalue().encode("utf-8-sig")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=fortress_master.csv"},
    )


@router.get("/master/xlsx")
async def export_master_xlsx(request: Request):
    """Download all SCRAPED companies across all queries as XLSX. Admin only."""
    user = getattr(request.state, 'user', None)
    if not user or user.role != 'admin':
        from fastapi.responses import JSONResponse
        return JSONResponse(status_code=403, content={"error": "Admin uniquement"})
    rows = await fetch_all(f"""
        WITH {merged_contacts_cte('SELECT DISTINCT siren FROM batch_tags')}
        SELECT {_EXPORT_SELECT}
        FROM (SELECT DISTINCT siren FROM batch_tags) qt
        JOIN companies co ON co.siren = qt.siren
        {_EXPORT_JOINS}
        WHERE (co.siren NOT LIKE 'MAPS%%' OR co.link_confidence IN ('confirmed', 'pending'))
          AND (co.naf_status IS DISTINCT FROM 'mismatch' OR co.link_method IN ('chain', 'gemini_judge', 'siret_address_naf'))
        ORDER BY (CASE WHEN co.link_confidence = 'pending' THEN 1 ELSE 0 END),
                 co.denomination
    """)
    return _to_xlsx(rows, "fortress_master.xlsx")


# ── FILTERED CONTACTS EXPORT (must be declared BEFORE {batch_id} routes) ──


@router.get("/contacts/csv")
async def export_contacts_filtered_csv(
    request: Request,
    q: str = None,
    department: str = None,
    naf_code: str = None,
):
    """Export contacts matching the current filters as CSV.

    Uses the same filter logic as /api/contacts/list. Returns ALL matching
    contacts (capped at 10,000), not just the currently loaded page.
    """
    user = getattr(request.state, "user", None)

    where_parts: list[str] = []
    params: list = []

    # Workspace filter — admin sees all, others scoped to their workspace
    if user and not user.is_admin:
        where_parts.append("qt.workspace_id = %s")
        params.append(user.workspace_id)

    if q:
        clean_q = q.strip()
        clean_digits = clean_q.replace(" ", "")
        if clean_digits.isdigit():
            where_parts.append("(co.siren LIKE %s)")
            params.append(f"{clean_digits}%")
        elif clean_q.upper().startswith("MAPS"):
            where_parts.append("(co.siren ILIKE %s)")
            params.append(f"{clean_q}%")
        elif len(clean_q) <= 2:
            where_parts.append("co.denomination LIKE %s")
            params.append(f"{clean_q.upper()}%")
        else:
            # Multi-field search — must match contacts_list.py exactly
            where_parts.append("""(
                co.denomination ILIKE %s
                OR co.enseigne ILIKE %s
                OR co.siren ILIKE %s
                OR co.ville ILIKE %s
                OR co.naf_code ILIKE %s
                OR co.departement = %s
                OR mc.phone ILIKE %s
                OR mc.email ILIKE %s
                OR mc.website ILIKE %s
                OR o.nom ILIKE %s
                OR o.prenom ILIKE %s
            )""")
            like = f"%{clean_q}%"
            dept_q = clean_q.strip()
            params.extend([like, like, like, like, like, dept_q, like, like, like, like, like])

    if department:
        where_parts.append("co.departement = %s")
        params.append(department.strip())

    if naf_code:
        where_parts.append("co.naf_code LIKE %s")
        params.append(f"{naf_code.strip().upper()}%")

    # Include confirmed AND pending MAPS rows — exclude only fully unmatched MAPS.
    # Pending rows are flagged via the `Statut lien` column and sorted to the bottom
    # so Cindy sees confirmed contacts first.
    where_parts.append("(co.siren NOT LIKE 'MAPS%%' OR co.link_confidence IN ('confirmed', 'pending'))")
    # Exclude NAF-mismatch rows — except chain, gemini_judge, and siret_address_naf matches which are included (high-confidence)
    where_parts.append("(co.naf_status IS DISTINCT FROM 'mismatch' OR co.link_method IN ('chain', 'gemini_judge', 'siret_address_naf'))")

    where_clause = " AND ".join(where_parts) if where_parts else "TRUE"

    # Cap at 10k rows — safety net, prevents massive exports.
    # DISTINCT ON uses COALESCE(linked_siren, co.siren) as the dedup key:
    # - MAPS entities linked to a real SIREN (confirmed OR pending) collapse into one row
    # - Unmatched MAPS entities (linked_siren IS NULL) stay as distinct rows (via siren)
    # - Pure SIRENE entities dedup by their own siren
    rows = await fetch_all(f"""
        WITH {merged_contacts_cte('SELECT DISTINCT qt2.siren FROM batch_tags qt2')}
        SELECT DISTINCT ON (COALESCE(co.linked_siren, co.siren)) {_EXPORT_SELECT}
        FROM batch_tags qt
        JOIN companies co ON co.siren = qt.siren
        LEFT JOIN merged_contacts mc ON mc.siren = co.siren
        LEFT JOIN LATERAL (
            SELECT * FROM officers off2 WHERE off2.siren = co.siren
            ORDER BY (off2.ligne_directe IS NOT NULL)::int DESC
            LIMIT 1
        ) o ON true
        LEFT JOIN (
            SELECT siren, STRING_AGG(text, ' | ' ORDER BY created_at DESC) AS notes
            FROM company_notes
            GROUP BY siren
        ) cn ON cn.siren = co.siren
        WHERE {where_clause}
        ORDER BY COALESCE(co.linked_siren, co.siren),
                 (CASE WHEN co.link_confidence = 'pending' THEN 1 ELSE 0 END),
                 co.denomination
        LIMIT 10000
    """, tuple(params) if params else None)

    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")
    writer.writerow([col[0] for col in _CSV_COLUMNS])
    for row in (rows or []):
        writer.writerow([_fmt(col[1], row.get(col[1])) for col in _CSV_COLUMNS])

    content = buf.getvalue().encode("utf-8-sig")

    # Build descriptive filename from filters
    filename_parts = ["contacts"]
    if department:
        filename_parts.append(f"dept{department}")
    if naf_code:
        filename_parts.append(f"naf{naf_code}")
    if q:
        filename_parts.append(q.strip()[:20].replace(" ", "_"))
    filename = "_".join(filename_parts) + ".csv"

    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── PER-QUERY EXPORTS ────────────────────────────────────────────


@router.get("/{batch_id}/csv")
async def export_csv(batch_id: str, request: Request, search_query: str = Query("", description="E4.A drill-down filter")):
    """Download all companies for a query as CSV."""
    user = getattr(request.state, "user", None)
    if user and not user.is_admin:
        batch = await fetch_one(
            "SELECT workspace_id FROM batch_data WHERE batch_id = %s", (batch_id,)
        )
        if not batch or batch["workspace_id"] != user.workspace_id:
            from fastapi.responses import JSONResponse
            return JSONResponse(status_code=403, content={"error": "Accès refusé."})
    rows = await _fetch_export_data(batch_id, search_query=search_query or None)

    # Build filename — include slugged search_query when drill-down is active (E4.A)
    if search_query:
        sq_slug = re.sub(r'[^a-zA-Z0-9]+', '_', search_query).strip('_')[:50]
        filename = f"{batch_id}_{sq_slug}.csv"
    else:
        filename = f"{batch_id}.csv"

    if not rows:
        return StreamingResponse(
            io.BytesIO(b"No data"),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")
    writer.writerow([col[0] for col in _CSV_COLUMNS])
    for row in rows:
        writer.writerow([
            _fmt(col[1], row.get(col[1])) for col in _CSV_COLUMNS
        ])

    content = buf.getvalue().encode("utf-8-sig")  # BOM for Excel
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/{batch_id}/jsonl")
async def export_jsonl(batch_id: str, request: Request):
    """Download all companies for a query as JSONL."""
    user = getattr(request.state, "user", None)
    if user and not user.is_admin:
        batch = await fetch_one(
            "SELECT workspace_id FROM batch_data WHERE batch_id = %s", (batch_id,)
        )
        if not batch or batch["workspace_id"] != user.workspace_id:
            from fastapi.responses import JSONResponse
            return JSONResponse(status_code=403, content={"error": "Accès refusé."})
    rows = await _fetch_export_data(batch_id)
    lines = []
    for row in rows:
        # Convert date objects to strings
        clean = {}
        for k, v in row.items():
            if hasattr(v, "isoformat"):
                clean[k] = v.isoformat()
            else:
                clean[k] = v
        lines.append(json.dumps(clean, cls=_DecimalEncoder, ensure_ascii=False))

    content = "\n".join(lines).encode("utf-8")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/jsonl",
        headers={"Content-Disposition": f"attachment; filename={batch_id}.jsonl"},
    )


@router.get("/{batch_id}/xlsx")
async def export_xlsx(batch_id: str, request: Request):
    """Download all companies for a query as XLSX."""
    user = getattr(request.state, "user", None)
    if user and not user.is_admin:
        batch = await fetch_one(
            "SELECT workspace_id FROM batch_data WHERE batch_id = %s", (batch_id,)
        )
        if not batch or batch["workspace_id"] != user.workspace_id:
            from fastapi.responses import JSONResponse
            return JSONResponse(status_code=403, content={"error": "Accès refusé."})
    rows = await _fetch_export_data(batch_id)
    return _to_xlsx(rows, f"{batch_id}.xlsx")


# ── BULK EXPORT ──────────────────────────────────────────────────


class BulkExportRequest(BaseModel):
    sirens: list[str]


@router.post("/bulk/csv")
async def export_bulk_csv(body: BulkExportRequest, request: Request):
    """Export selected SIRENs as CSV (from dashboard bulk selection)."""
    if not body.sirens:
        return StreamingResponse(
            io.BytesIO(b"No data"),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=fortress_selection.csv"},
        )

    user = getattr(request.state, "user", None)
    if user and not user.is_admin:
        # Scope: only export SIRENs that belong to the user's workspace via batch_tags
        rows = await fetch_all(f"""
            WITH workspace_sirens AS (
                SELECT DISTINCT bt.siren
                FROM batch_tags bt
                JOIN batch_data bd ON bd.batch_id = bt.batch_id
                WHERE bd.workspace_id = %s AND bt.siren = ANY(%s)
            ),
            {merged_contacts_cte('SELECT siren FROM workspace_sirens')}
            SELECT {_EXPORT_SELECT}
            FROM workspace_sirens ws
            JOIN companies co ON co.siren = ws.siren
            {_EXPORT_JOINS}
            WHERE (co.siren NOT LIKE 'MAPS%%' OR co.link_confidence IN ('confirmed', 'pending'))
              AND (co.naf_status IS DISTINCT FROM 'mismatch' OR co.link_method IN ('chain', 'gemini_judge', 'siret_address_naf'))
            ORDER BY (CASE WHEN co.link_confidence = 'pending' THEN 1 ELSE 0 END),
                     co.denomination
        """, (user.workspace_id, body.sirens))
    else:
        rows = await fetch_all(f"""
            WITH {merged_contacts_cte('SELECT UNNEST(%s::text[])')}
            SELECT {_EXPORT_SELECT}
            FROM companies co
            {_EXPORT_JOINS}
            WHERE co.siren = ANY(%s)
              AND (co.siren NOT LIKE 'MAPS%%' OR co.link_confidence IN ('confirmed', 'pending'))
              AND (co.naf_status IS DISTINCT FROM 'mismatch' OR co.link_method IN ('chain', 'gemini_judge', 'siret_address_naf'))
            ORDER BY (CASE WHEN co.link_confidence = 'pending' THEN 1 ELSE 0 END),
                     co.denomination
        """, (body.sirens, body.sirens))

    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")
    writer.writerow([col[0] for col in _CSV_COLUMNS])
    for row in (rows or []):
        writer.writerow([_fmt(col[1], row.get(col[1])) for col in _CSV_COLUMNS])

    content = buf.getvalue().encode("utf-8-sig")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=fortress_selection.csv"},
    )


# ── XLSX helper ──────────────────────────────────────────────────


def _to_xlsx(rows: list[dict] | None, filename: str) -> StreamingResponse:
    """Convert rows to an XLSX file and return as StreamingResponse.
    Pending rows (link_confidence='pending') are filled with soft peach
    (FFE6CC) so they're visually distinct from confirmed rows."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Export"

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a3e", end_color="1a1a3e", fill_type="solid")
    for col_idx, (label, _) in enumerate(_CSV_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Pending-row fill: soft peach matching the --warning palette in the UI
    pending_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")

    # Data rows
    for row_idx, row in enumerate(rows or [], 2):
        is_pending = (row.get("link_confidence") == "pending")
        for col_idx, (_, key) in enumerate(_CSV_COLUMNS, 1):
            val = row.get(key)
            if hasattr(val, "isoformat"):
                val = val.isoformat()
            cell = ws.cell(row=row_idx, column=col_idx, value=_fmt(key, val))
            if is_pending:
                cell.fill = pending_fill

    # Auto-width (approximate)
    for col_idx, (label, _) in enumerate(_CSV_COLUMNS, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(len(label) + 4, 12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
